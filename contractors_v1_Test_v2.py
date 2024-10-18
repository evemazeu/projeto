# -*- coding: utf-8 -*-
"""
Created on Thu Mar  5 16:36:14 2020
@author: MINORGOMEZVIALES
@autor: Edinelson Almeida
"""

import pandas as pd
import numpy as np
import requests
import os
import urllib3
import datetime
import pprint
import time
import concurrent.futures
from datetime import timedelta
from datetime import date
import openpyxl as xl
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import threading
import logging, logging.handlers
import configparser
from configparser import ConfigParser, ExtendedInterpolation
import ast
from tqdm import tqdm
import json as js
from warnings import simplefilter

#added to don't show alerts of poor performance
simplefilter(action="ignore", category=pd.errors.PerformanceWarning)

#%%  
pp = pprint.PrettyPrinter(indent=4)
urllib3.disable_warnings()
#%% 
#Global variables
try:    
    config_normal = configparser.ConfigParser()
    configInterpolation = ConfigParser(interpolation=ExtendedInterpolation())
    configInterpolation.read('configContractors_v1.ini')
    config_normal.read('configContractors_v1.ini')
    
    workloc = config_normal.get('basic_params','workloc')
    filename_cld        = config_normal.get('basic_params','filename_cld')    
    #variable required in process_AvailMatch() and process_df_to_excel() functions
    today = datetime.date.fromtimestamp(time.time())
    # today   = datetime.datetime.now().isoformat()[0:10]   
    # df_final = pd.DataFrame()
    
   #BEGIN log configuration
    LOG_FILENAME = "logs\\contractors_" + str(datetime.datetime.utcnow().strftime('%Y_%m_%d_%I_%M_%S')) + '.log'    
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)   
    formatter = logging.Formatter('%(asctime)s:%(levelname)s: %(message)s')
    # stream_formatter = logging.Formatter('%(levelname)s: %(message)s')    
    file_handler = logging.FileHandler(LOG_FILENAME)
    file_handler.setFormatter(formatter)   
    stream_handler = logging.StreamHandler()
    # stream_handler.setFormatter(stream_formatter)   
    logger.addHandler(file_handler)
    # logger.addHandler(stream_handler)
   #END log configuration
    
except configparser.ParsingError as e:
    print("Error reading the file:")
    print(e)
    logger.error('ParsingError loading config file:', exc_info=True)
except configparser.NoSectionError as e:
    print("Specified section is not found:")
    print(e)
    logger.error('NoSectionError loading config file:', exc_info=True)
except configparser.NoOptionError as e:
    print("Specified option is not found in the specified section:")
    print(e)
    logger.error('NoOptionError loading config file:', exc_info=True)    
except configparser.Error as e:
    print("Error: ")
    print(e)
    logger.error('Exception loading config file:', exc_info=True)

#%%
def createNoMatches_df():
    """
    creates empty columns dataframe required in process_API_data() function
    
    Returns
    ----------
    dataFrame
        a dataframe with columns names only
    """
    try:
        empty = pd.DataFrame()
        noMatches_df = pd.DataFrame()
        PractitionerColumns = ast.literal_eval(config_normal.get('colsConfig','practitionercolumns'))    
        seat_fields_to_retrieve = [x for x in PractitionerColumns]   
        #creates empty columns dataframe required in process_API_data() function
        noMatches_df = pd.DataFrame(empty, columns = seat_fields_to_retrieve)  
        noMatches_df.fillna('None', inplace = True)  
        return noMatches_df

    except Exception:
          logger.error('Exception at createNoMatches_df():', exc_info=True)
#%% 
def pmp_imt(row):
    """
        a function to classify 'mysaWorkLocationIMT' for demand dataFrame 
    
    Parameters
    -----------
    row: str
        Receives column 'ORG_LVL_3_TXT' from demand df
    Returns
    --------
    str
        either values from: 'ORG_LVL_3_TXT' or 'ORG_LVL_4_TXT'
    """
    try:
        
        if row['ORG_LVL_3_TXT'] in ('LA','MEA','GCG'):
            return row['ORG_LVL_4_TXT']
        else:
            return row['ORG_LVL_3_TXT'] 
    except Exception:
            logger.error('Exception at pmp_imt():', exc_info=True)
            
#%%
def date_to_iso(row):
    """
        a function to convert 'endDate' to iso format 
    
    Parameters
    -----------
    row: date
        Receives date column 'SUBK_ACTL_END_DT' from demand df
    Returns
    --------
    date
        'SUBK_ACTL_END_DT' transformed as iso date 
    """
    
    try:
        row['SUBK_ACTL_END_DT'] = pd.to_datetime(row['SUBK_ACTL_END_DT'])
        return row['SUBK_ACTL_END_DT'].isoformat()
    except Exception as ex:
            print(ex)
            logger.error('Exception date_to_iso:', exc_info=True) 
#%%
def reduce_matchScore(mmetrics_val):
    """
    a function to calculate the reduction for match_score 
    
    Parameters
    -----------
    mmetrics_val: str
        Receives P_Matching_Metrics as string
    Returns
    --------
    list    
        total: int
            the total points to reduce in field 'P_matchScore'
        cell_val: str
            The new string for 'P_Matching_Metrics' field 
    """
    
    try:

        items_to_reduce =  {"AVAIL_DATE_WITHIN_14_DAYS_OF_START_DATE":(5.0),
               "AVAIL_DATE_WITHIN_7_DAYS_OF_START_DATE":(10.0) 
               # ,"REQUIRED_SKILLS_MATCH_100": (30.0)
               # ,"LANGUAGES_MATCH_FULL": (10.0)
               }
        delimiter = ','
        
        cell_val = delimiter.join(mmetrics_val)
        total = 0

        for key, value in items_to_reduce.items():
            if key in cell_val:
                total = total + value
                text_to_find = key+" ("+str(value) + ")"
                text_to_replace = key+" (0.0)"
                cell_val = cell_val.replace(text_to_find, text_to_replace)
    
        return [total,cell_val]
    
    except Exception:
            logger.error('Exception at reduce_matchScore:', exc_info=True)    
#%%
def create_proximity_date(todaysDate, dayToFix):
    """
        a function to get weekday's date 
    
    Parameters
    -----------
    todaysDate: date
        Receives today's date
    dayToFix: str 
        values as "monday, tuesday, etc..."
    Returns
    --------
    date
        date for weekday in the future.
    """
    try:
        weekdays = {'monday':1, 'tuesday':2, 'wednesday':3, 'thursday':4, 'friday':5,'saturday':6,'sunday':7}
        dayToFix = weekdays[dayToFix.lower()]  
        current_date_val = todaysDate.isoweekday()        
        difference = (dayToFix - current_date_val)+7    
        newDate = (todaysDate + timedelta(days=int(difference))).isoformat()[0:10] 
    
        return newDate
    except Exception:
            logger.error('Exception in create_proximity_date():', exc_info=True)     
#%% 
def create_demand(filename_cld, worklocVal):
    """ 
        a function to form demand dataframe from cld file
    
    Parameters
    ----------
    filename_cld: dataframe
        a valid cld dataframe
    
    Returns
    -----------
    dataframe
        a dataframe with cld data
    """
    try:

        lob                 = config_normal.get('basic_params','lob')
        partition           = config_normal.get('basic_params','partition')        
        
        #Get the file data
        demand_df = pd.read_csv(filename_cld, index_col = 'CNUM_ID')   
        #Creating artificial fields for Open Seats matching rules    
        demand_df['mySAIndicator']           = '1'
        demand_df['mySASeatLob']             = lob # Possible values are GBS (includes GPS), GTS, Security
        
        
        if worklocVal == ['Greater China Group']:
            demand_df['type'] = np.where(demand_df['GD_FLG'] == 'Y','GIC open seat','Geo open seat') # Possible values are "Geo open seat" and "GIC open seat"
        else:    
            demand_df['type'] = 'Geo open seat'   #We expect only "GD_FLG=N" for all Markets except GCG. Should pull Domestic only via rule set in DataStage. 
                 
            
        demand_df['startDate']               = create_proximity_date(today, "Monday")  +' 04:00:00 GMT'
        # demand_df['startDate']             = today+' 04:00:00 GMT' #should be next monday 
        demand_df['endDate']                 = demand_df.apply(date_to_iso,axis=1) +' 04:00:00 GMT'
        demand_df['requestedBandLow']        = demand_df['CNTRCTR_EQUIV_CD'].astype(str)
        demand_df['requestedBandHigh']       = demand_df['CNTRCTR_EQUIV_CD'].astype(str)
        demand_df['mysaWorkLocationIMT']     = demand_df.apply(pmp_imt,axis=1)
        demand_df['partition']               = partition
        demand_df.rename(columns ={'CTRY_NM':'mysaWorkLocationCountry'}, inplace = True)
        # demand_df.rename(columns ={'ORG_LVL_1_TXT':'mysaWorkLocationIOT'}, inplace = True)
        demand_df.rename(columns ={'SKL_SET_NM':'skillSet'}, inplace = True)
        demand_df.rename(columns ={'JOB_ROLE_NM':'jobRole'}, inplace = True)
        
        demand_df['niceToHaveSkills']   = demand_df['REQ_SKL_TXT']
        demand_df['requiredSkills']     = demand_df['REQ_SKL_TXT']
        demand_df.rename(columns={'LANGUAGE':'requestedLanguage'}, inplace = True)
    
       
        demand_df.fillna('', inplace = True)  
        
        return demand_df
    except Exception:
            logger.error('Exception in create_proximity_date():', exc_info=True)
            
#%%
def createPayload(worklocVal):
    """ 
        a function for open seats, for the matching rules to be added in wex-api
    
    Parameters
    ----------
    worklocVal: str
        a valid worklocation value
    
    Returns
    -----------
    dict
        a dictionary with all the matching rules.
    """
    
    try:
        #basic parameters
        records             = config_normal.get('basic_params','records')
        topnrecords         = config_normal.get('basic_params','topnrecords')
        availabilitydays    = config_normal.get('basic_params','availabilitydays')
        # language            = config_normal.get('basic_params','language')
       
        worklocation_IOT = ""
        worklocation_IMT = ""

        if worklocVal == ["North America"]:
            worklocation_IOT = "Americas"
            worklocation_IMT = "! LA"
            
        if worklocVal == ["Latin America"]:
             worklocation_IOT = "Americas"
             worklocation_IMT = "LA"
             
        if worklocVal == ["APAC"]:
            worklocation_IOT = "APAC"   
            worklocation_IMT = "! GCG"
            
        if worklocVal == ["Greater China Group"]:
            worklocation_IOT = "APAC"   
            worklocation_IMT = "GCG"
            
        if worklocVal == ["MEA"]:
            worklocation_IOT = "EMEA"
            worklocation_IMT = "MEA"
            
        if worklocVal == ["Europe"]:
            worklocation_IOT = "EMEA"
            worklocation_IMT = "! MEA"
            
        if worklocVal == ["Japan"]:
            worklocation_IOT = "JP"
            worklocation_IMT = "Japan"
            
        matchingRules_payload = {
          "numberOfRequestedRecords": records,
          "numberOfRequestedRecordsPostSorting": topnrecords,
          "configurations": {
            "*": {
              "name": "*",
              "description": "Configurations in this partition will override all the others",
              "matchingRules": {"matchBand": True,
                                "matchAvailabilityDate": True,
                                "matchPrimaryJrs": True,
                                "matchWorkLocation":  False,
                                "matchLob": True,
                                "matchNiceToHaveSkills": True,
                                "matchMustHaveSkills": True,
                                "matchLanguage": True,
                                "matchResourceType": True,
                                "matchSecondaryJrs": True},
             "rankingRules": { "sortOrder": ["MATCH_SCORE"],
                              "matchScoreConfiguration": {
	   		                    "quantifiers":["BAND_SCORE","AVAILABILITY_SCORE","JRS_SCORE","LANGUAGE_SCORE", "LOCATION_SCORE", "REQUIRED_SKILLS_SCORE"],
                    	   		"languageQuantifiers": {
			                            "LANGUAGES_MATCH_FULL": 10.0,
			                            "LANGUAGES_MATCH_PARTIAL": 5.0,
			                            "NO_LANGUAGES_MATCH": 0.0,
			                            "LANGUAGES_NOT_REQUIRED": 0
                                        } #End of languageQuantifiers                                  
                                   ,"bandQuantifiers":{
                                              "BAND_WITHIN_RANGE": 10.0,
                                              "BAND_BELOW_1_RANGE": 5.0,
                                              "BAND_ABOVE_1_RANGE": 3.0,
                                              "NO_BAND_MATCH": 0.0 
                                              } # End of bandQuantifiers
                                   ,"requiredSkillsQuantifiers": {
                                                            "REQUIRED_SKILLS_MATCH_100": 30.0,
                                                            "REQUIRED_SKILLS_MATCH_90": 27.0,
                                                            "REQUIRED_SKILLS_MATCH_80": 24.0,
                                                            "REQUIRED_SKILLS_MATCH_70": 21.0,
                                                            "REQUIRED_SKILLS_MATCH_60": 18.0,
                                                            "REQUIRED_SKILLS_MATCH_50": 15.0,
                                                            "REQUIRED_SKILLS_MATCH_40": 12.0,
                                                            "REQUIRED_SKILLS_MATCH_30": 9.0,
                                                            "REQUIRED_SKILLS_MATCH_20": 6.0,
                                                            "REQUIRED_SKILLS_MATCH_10": 3.0,
                                                            "NO_REQUIRED_SKILLS_MATCH": 0.0,
                                                            "REQUIRED_SKILLS_NOT_PROVIDED": 0.0
                                                            } #End of Required SkillsQuantifiers
                                   } # end of matchScoreConfiguration
                              }# en of rankingRules
             ,"immediateAvailabilityDays": availabilitydays,
             "lowBandSlack": 1, 
             "highBandSlack": 1,
             "maxMustHaveKeywords": 0,
             "maxNiceToHaveKeywords":-1,
            "fieldsToRetrieve": ['AVAILABILITY_DATE'
                              ,'BAND'
                              ,'CNUM_ID'
                              ,'FULL_NAME'
                              ,'GEO_SPECIFIC_ORGANIZATION'
                              ,'INDUSTRY'
                              ,'INDUSTRY_EXPERTISE'
                              ,'LANGUAGE'
                              ,'MYSA_TALENT_POOL_LOB'
                              ,'NOTES_ID'
                              ,'PRIMARY_JRSS'
                              ,'RDM_COMMENTS'
                              ,'RESOURCE_TYPE'
                              ,'SECONDARY_JRSS'
                              ,'SERVICE'
                              ,'WORK_LOCATION_CITY'
                              ,'WORK_LOCATION_COUNTRY'
                              ,'WORK_LOCATION_IMT'
                              ,'WORK_LOCATION_IOT'
                              ,'SERVICE_AREA'
                              ,'SECTOR'
                              ,'RSA_VALUES_MASK']}
          },
          "targetDocumentType": "PRACTITIONER",
          "targetDocumentSelection":[
              {"name": "WORK_LOCATION_IOT",  'selections': [worklocation_IOT]},
              {"name": "WORK_LOCATION_IMT",  'selections': [worklocation_IMT]},
              {"name": "DEPLOYABILITY", "selections": ["Full"]},
              {"name": "WORK_LOCATION_COUNTRY", "selections": ["! Austria"]},
              {"name": "XPATH", "selections":["xpath not( $WORK_LOCATION_COUNTRY='United States' and ($DIVISION='16' or $DIVISION='1P') )"]}],
         
          }#End of Payload

        
        return matchingRules_payload
    
          # "facetSelections":[{"name": "WORK_LOCATION_IOT", "selections":[worklocation]}],
          # "xPathFilter" : ["not($WORK_LOCATION_COUNTRY = 'Austria')"]    
        #{"name":"LANGUAGE_FACET","selections":[language]},
         
      
    except Exception:
        logger.error('Exception createPayload():', exc_info=True)

#%%
def process_AvailMatch(field):
    """ 
        a function to classify bench for Available Match
    
    Parameters
    ----------
    field: date
        P_availabilityDate field from practitioners dataFrame
    
    Returns
    -----------
    str
        a string value to be added in Avail Match field.
    """
    
    try:
        day = int(field.dt.day)
        month = int(field.dt.month)
        year = int(field.dt.year)
        dateReceived =  date(year, month, day).isoformat()
        
        today   = datetime.datetime.now().isoformat()[0:10]    
        d7  = (datetime.datetime.now() + timedelta(days=int(7))).isoformat()[0:10]
        d14 = (datetime.datetime.now() + timedelta(days=int(14))).isoformat()[0:10]
        d21 = (datetime.datetime.now() + timedelta(days=int(21))).isoformat()[0:10]
        d30 = (datetime.datetime.now() + timedelta(days=int(30))).isoformat()[0:10]
        
        if dateReceived< today:
            match = "On Bench"
        elif dateReceived <= d7:
           match =   "Bench in 1-7 days"  
        elif dateReceived <= d14:
           match =  "Bench in 8-14 days"
        elif dateReceived <= d21:
           match =  "Bench in 15-21 days"
        elif dateReceived <= d30:
           match = "Bench in 22-30 days"
        else:
           match =  "Bench in 31-56 days"

        return match
                
    except Exception:
        # print("Error: ")
        logger.error('Exception process_API_data:', exc_info=True)
#%%
def get_API_Data(OsId, payload, demand_df):
    """
        a function to get data from wex api
    
    Parameters
    ------------
    OsId: str
        Serial number
    payload: dictionary
        matching rules dictionary
    demand_df: DataFrame
        cld dataframe from csv file.
    
    returns: 
    ---------
        JSON
        a json structure for for openSeats
    """
    #API related parameters
    pw    =  os.getenv('INET_PW') # your intranet id password
    uid   =  os.getenv('INET_UID') # your intranet id password
        
    url      = config_normal.get('api','url')
    status_dict = dict()
    
    
    
    try:
        payload['openSeat'] = demand_df.loc[OsId,['mySAIndicator','mySASeatLob','type','jobRole','skillSet','startDate',
                                             'endDate','requestedBandLow','requestedBandHigh','mysaWorkLocationIMT',
                                             'mysaWorkLocationCountry','niceToHaveSkills','requiredSkills','requestedLanguage']].to_dict()        

        language = demand_df.loc[OsId,['requestedLanguage']].to_list() 
        payload['openSeat'].update({'requestedLanguage': language})

        with open('payload_'+OsId+'.json', 'w', encoding='utf-8') as f:
            js.dump(payload, f,ensure_ascii=False, indent=4)
       
        logger.debug(payload)

        #Performs connection to the API
        r = requests.post(url,auth=(uid, pw),verify=False,json=payload)
        status = r.status_code
        c = r.json()

#        with open('output_'+OsId+'.json', 'w', encoding='utf-8') as f:
#             js.dump(c, f,ensure_ascii=False, indent=4)

      
        if len(r.text) == 0:
            logger.debug("Blank response requests error")
        
        if r.status_code == 200:
            c['status_code'] = status
#            logger.info('\tStatus:{}:\tTotMatches:{}\tOSid:{}'.format(r.status_code,c['totalMatches'],OsId))
            return c     
       
        if r.status_code == 400:
            """In other words, the data stream sent by the client to the server didn't follow the rules."""          
            logger.error("400 Bad Request, invalid sintaxis for: {}".format(OsId))
            status_dict['status_code'] = 400
            return status_dict
        
        if r.status_code == 500:
            logger.debug("Internal server error for {}\n".format(OsId)) 
            status_dict['status_code'] = 500
            return status_dict
            
    except requests.exceptions.RequestException as e: 
            # print('***Exception***')
            print(e)
            logger.error('RequestException get_API_Data:', exc_info=True)
    except Exception as ex:
            print(ex)
            logger.error('Exception get_API_Data:', exc_info=True)
    # finally:
    #     logger.info('Status:{}: TotMatches:{} WorkLoc:{} OSid:{}'.format(r.status_code,c['totalMatches'],payload['openSeat']['mysaWorkLocationIOT'],OsId))
    
#%%
def process_API_data(fake_OsId, json, demand, Pract_NoMatches_df):
    """
        A function to ensemble raw data from API and CLD csv data file
   
    Parameters
   ----------
   fake_OsId: str
       Serial number from CLD csv file
   json: dictionary
       API formated JSON data
   demand: DataFrame
       A dataframe created from CLD file
   Pract_NoMatches_df: DataFrame
       a columns only empty dataframe to be used for practitioners with no matches
   
   Returns
   ---------
   Dataframe
       a raw dataframe required by format_final_df() function 
    """

    #global df_final
    df_final = pd.DataFrame()
    temp = pd.DataFrame() #local variable to merge and concatenate dataframes
    temp2 = pd.DataFrame() #local variable to set top N qualityScore_OS
    practs_counter = 0
    try:
        if json['status_code'] == 200:
            
            #OpenSeats data
            OpenSeat_df = pd.DataFrame(data=[json['openSeat'].values()],columns=json['openSeat'].keys()) 
            
            #create columns based from demand_df
            OpenSeat_df['osID']                         = fake_OsId
            OpenSeat_df['Contractor']                   = demand.loc[fake_OsId,'PRSN_NM']
            OpenSeat_df['Manager_']                     = demand.loc[fake_OsId,'MGR_NOTES_EMAIL_ID']    
            OpenSeat_df['Client_']                      = demand.loc[fake_OsId,'PO_CLIENT_NM']
            OpenSeat_df['Project_']                     = demand.loc[fake_OsId,'PO_PROJECT_NM']
            OpenSeat_df['Subcontractor_Category_']      = demand.loc[fake_OsId,'SUBCONTRACTOR_CAT']
            OpenSeat_df['Contractor Start Date']        = demand.loc[fake_OsId,'CLD_CREATE_DT'] 
            OpenSeat_df['Required Language']            = demand.loc[fake_OsId,'requestedLanguage']  
            OpenSeat_df['OpenSeatId']                   = demand.loc[fake_OsId,'OPEN_SEAT_ID']  
            OpenSeat_df['Tram_Request_Id']              = demand.loc[fake_OsId,'TRAM_REQUEST_ID']
            OpenSeat_df['workLocationIOT']              = demand.loc[fake_OsId,'ORG_LVL_1_TXT']
            OpenSeat_df['PO_end_date']                  = demand.loc[fake_OsId,'PO_END_DATE']
   
            #extract practitioners
            Practitioner_df = json['practitioners']  
            
            #Apply suffix '_P' to all column names
            OpenSeat_df = OpenSeat_df.add_prefix('D_')
            #Creates a dummy column for matching on merge process
            OpenSeat_df['common_key'] = 0  
           
            #Loops thru each seats found in Seats_df dataframe
            for id in Practitioner_df:
                practs_counter +=1
                #Creates OSeats temporal dataFrame with Seat Information
                Practitioner_temp = pd.DataFrame(data=[id.values()],columns=id.keys())                
                #converts 'availabilityDate' to date format
                Practitioner_temp['availabilityDate'] = pd.to_datetime(Practitioner_temp['availabilityDate'])                          
                #Apply suffix '_OS' to all column names
                Practitioner_temp = Practitioner_temp.add_prefix('P_')       
                #Creates 'AvailMatch_M' with bench match, calls function
                Practitioner_temp['Avail Match'] =  process_AvailMatch(pd.to_datetime(Practitioner_temp.P_availabilityDate))               
                #Creates 'QualityScoreQuantifiers_OS' column, converts list inside cell to String
                # Practitioner_temp['P_QualityScoreQuantifiers'] = ', '.join(map(str, id['qualityScoreQuantifiers']))
                Practitioner_temp['P_Matching_Metrics'] = ', '.join(map(str, Practitioner_temp['P_matchScoreQuantifiers']))                
                #Creates a dummy column for matching on merge process
                Practitioner_temp['common_key'] = OpenSeat_df['common_key']              
                #Creates column with the number of matches from json API
                Practitioner_temp['P_TotalMatches'] = json['totalMatches']               
                #Converts 'qualityScore_OS' as float
                Practitioner_temp['P_matchScore'] =  Practitioner_temp['P_matchScore'].astype("float")
                
                
                #To activate the reduction, uncomment these lines
                # listToReduce = reduce_matchScore(Practitioner_temp['P_Matching_Metrics'])                
                # Practitioner_temp['P_matchScore'] = Practitioner_temp['P_matchScore'] - listToReduce[0]
                # Practitioner_temp['P_Matching_Metrics'] = listToReduce[1]
                
                #Merge supply with Oseats using dummy key 'common_key', join type 'outer'
                temp2 = pd.merge(OpenSeat_df, Practitioner_temp, on='common_key', how ='outer')               
                #concatenate DataFrame to df_final
                df_final = pd.concat([df_final,temp2]).reset_index(drop=True)                  
            #verifies if practitioner dataFrame is empty        
            if json['totalMatches'] == 0:
                #Creates a dummy column for matching on merge process
                Pract_NoMatches_df['common_key'] = OpenSeat_df['common_key']               
                #Creates a dummy column with the number of matches which is 0
                Pract_NoMatches_df['P_TotalMatches'] = 0
                Pract_NoMatches_df['Avail Match'] =  None
                Pract_NoMatches_df['P_matchScore'] =  0  
                Pract_NoMatches_df['P_Matching_Metrics'] =  None 
                #Merge supply with OS_NoMatches_df using dummy key 'common_key', join type 'outer'
                temp = pd.merge(OpenSeat_df, Pract_NoMatches_df, on='common_key', how ='outer')               
                #concatenates final dataframe with temp dataframe
                df_final = pd.concat([df_final,temp])
                                
            return df_final
        
        # else:
        #     return json
    
    except SyntaxError as e:
        # print('***Exception***')
        print(e)
        logger.error('SyntaxError process_API_data:', exc_info=True)
        
    except Exception as exception:
        # Output unexpected Exceptions.
        print(exception, False)
        logger.error('Exception process_API_data:', exc_info=True)
    
#%%
def format_final_df(dfVal):
   """ A function to format dataframe
   Parameters
   ----------
   dfVal: Dataframe
       a valid dataframe
       
   Returns
   ---------
   Dataframe
       a final dataframe required by process_df_to_excel() function 
   """

   #Columns settings
   #to literally pass in a list then you can use: ast.literal_eval()
   columnsToRename             = ast.literal_eval(config_normal.get('colsConfig','renamecolumns'))
   columnsToDrop               = ast.literal_eval(config_normal.get('colsConfig','dropcolumns'))
    
   df_final = dfVal 


    
   try:
       #checks df_final is valid to process
       if len(df_final)>0:
                #Step 1 Create index
                df_final.rename(columns ={'D_osID':'Contractor CNUM'}, inplace = True)
                df_final.set_index('Contractor CNUM', inplace = True)
                #Step 3 Create columns    
                df_final['D_JRSS'] =  df_final['D_jobRole']+'-'+df_final['D_skillSet']    
                df_final['P_JRSS'] =  df_final['P_jobRole'] +  df_final['P_skills']
                df_final['Template Version']= 1
                df_final['total_count']= 1
                df_final['Date of Data Retrieval']=datetime.datetime.now().isoformat()[0:10]
                #Step 4 Sort columns
                #sort 'P_cnum','jrsMatchType_OS' - ascending order
                df_final.sort_index(inplace=True)
                df_final.sort_values(by=['P_jrsMatchType'], inplace=True)    
                # sort qualityScore_OS - descending order
                df_final.sort_values(by=['P_matchScore'], inplace=True, ascending=False)

                # Step 5. Converting date columns to values: P_Avail_Date,D_Start_Date, D_End_Date.
                if  df_final['P_availabilityDate'] is None:
                     df_final['P_availabilityDate'] = pd.to_datetime(date.today() - timedelta(days=1))
                
                # Remove duplicate values from series
#                df_final = df_final.drop_duplicates(subset='P_availabilityDate', keep='first')
                df_final['P_availabilityDate'] = pd.to_datetime(df_final['P_availabilityDate'], utc=True)

                if  df_final['D_startDateFormated'] is None:
                     df_final['D_startDateFormated'] = create_proximity_date(today, "Monday")  +' 04:00:00 GMT'
                df_final['D_startDateFormated'] = pd.to_datetime(df_final['D_startDateFormated'], utc=True)
#                df_final['D_endDateFormated'] = pd.to_datetime(df_final['D_endDateFormated'])

                
                #Step 6. formatting the date columns to short dates
                df_final['P_availabilityDate'] = df_final['P_availabilityDate'].dt.date
                df_final['D_startDateFormated'] = df_final['D_startDateFormated'].dt.date

                # IllegalCharacterError - replace VT or STX by space for 'P_rdmComments' column
                df_final['P_rdmComments'] = df_final['P_rdmComments'].str.replace(r'[\u000b\u0002]', '\u0020', regex=True)

                df_final['P_rsaValues'] =  df_final.P_rsaValues.replace(np.nan,'None').astype('str')
                df_final['P_rsaValues'] =  df_final.P_rsaValues.replace("",'None').astype('str')
                if df_final.iloc[0]['P_rsaValues'] == "None":
                   df_final['P_rsaValues'] = 'None'
                else:    
                   test = df_final['P_rsaValues'].str.split("|", n = 5 , expand = True)    
                   df_final['P_rsaValues'] = test[2]+test[3]+" "+test[4]

                # create column for required skills percent
                df_final['requiredSkillsPercent'] = df_final['P_Matching_Metrics'].astype('str')
                df_final['requiredSkillsPercent'] = df_final['requiredSkillsPercent'].str.split(",",n=6, expand=True)[5]
                df_final = df_final.reset_index()

                for ind, row in df_final.iterrows():
                    if df_final.loc[ind, 'requiredSkillsPercent'] == ' NO_REQUIRED_SKILLS_MATCH (0.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '0'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_NOT_PROVIDED (0.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '0'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_10 (3.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '10'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_20 (6.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '20'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_30 (9.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '30'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_40 (12.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '40'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_50 (15.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '50'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_60 (18.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '60'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_70 (21.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '70'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_80 (24.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '80'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_90 (27.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '90'
                    elif df_final.loc[ind, 'requiredSkillsPercent'] == ' REQUIRED_SKILLS_MATCH_100 (30.0)':
                        df_final.loc[ind, 'requiredSkillsPercent'] = '100'
                    

                #Step 2 Drop columns from DataFrame
                df_final.drop(columnsToDrop, axis = 1, inplace = True)             
                #Step 8 rename columns
                df_final.rename(columns=columnsToRename, inplace = True)
                #Step 9 Fill.na cells
                df_final['Required Language'].replace('none', "Not Found", inplace = True)
                df_final.fillna('none', inplace = True)
                df_final.to_excel('finaldataframe.xlsx', encoding = 'utf-8', index = True)

                return df_final
       else:
           raise AttributeError('Dataframe is empty')
           logger.debug("Dataframe is empty")

       # return df_final
   
   except SyntaxError as e:
        # print('***Exception***')
        print(e)  
        logger.error('SyntaxError format_final_df:', exc_info=True)
   except Exception as exception:
        # Output unexpected Exceptions.
        print(exception, False)    
        logger.error('Exception format_final_df:', exc_info=True)    
    
#%%
def process_df_to_excel(worklocVal, finalDataFrame):
   """ A function to create excel report (final step)
   Parameters
   ----------
   worklocVal: str
       a valid worklocation ("North America", "Japan", "LatinAmerica", etc)
   finalDataFrame: DataFrame
       a valid Dataframe
   
   """
   try: 

        #Local variables
        # wrk_dir              = config_normal.get('excel','wrk_dir') 
        w46_template_fn      = config_normal.get('excel','w46_template') 
        W46_source_worksheet = config_normal.get('excel','w46_source_worksheet') 
        W46_about_worksheet  = config_normal.get('excel','w46_about_worksheet') 
        templateVersion      = config_normal.get('excel','templateVersion') 
        W46_saveTo          = config_normal.get('excel','w46_excel_saveTo')
        W46_template_location = config_normal.get('excel','w46_template_path')
        W46_saveToRaw         = config_normal.get('excel','wrk_dir_raw')
        nameOfExcelFile            =""
        
        df_final = pd.DataFrame()
        df_final = finalDataFrame
        
          

        if worklocVal == ["North America"]:
            nameOfExcelFile = "NA"

        if worklocVal == ["Latin America"]:
            nameOfExcelFile = "LA"

        if worklocVal == ["APAC"]:
            nameOfExcelFile = "AP"   

        if worklocVal == ["Greater China Group"]:
            nameOfExcelFile = "GCG"   
            
        if worklocVal == ["MEA"]:
            nameOfExcelFile = "MEA"
            
        if worklocVal == ["Europe"]:
            nameOfExcelFile = "EU"
            
        if worklocVal == ["Japan"]:
            nameOfExcelFile = "JP"
                   
       
        #change to working directory 
        # os.chdir(wrk_dir)       
        #Reset index
        df_final = df_final.reset_index()

        #df_final.reset_index(drop = True, inplace = True)  

        # Removes column 'index' as it only contains '0's
        if 'index' in df_final.columns:
            #source1_df = source1_df
            df_final.drop('index',axis = 1, inplace=True)        
        
        fn = W46_template_location + w46_template_fn
        
        #load_workbook in memory to open the template and assigned it to 'wb' variable
        wb = xl.load_workbook(filename = fn ,keep_vba=True)
        #Assigns 'about' worksheet to variable
        ws=wb[W46_about_worksheet]
        #assigns version to cell B3 in About Tab
        ws['B3']=templateVersion
        #assigns today's date to cell B4
        ws['B4']=today
        #Assigns 'Source1' worksheet to variable ('ws')
        ws=wb[W46_source_worksheet]
        #Removes old sheet 'Source1' from main template workbook (wb)
        wb.remove(ws)
        #Adds 'Source1' tab and set it as the active sheet
        ws = wb.create_sheet(W46_source_worksheet)
        #transfer the 'source1_df' to the active sheet using append(), removes index and keeps the headers
        for r in dataframe_to_rows(df_final, index=False, header=True):
            # logger.debug(f"\nexcel line:{r}\n")
            ws.append(r)       
        #Saves the file with new name and replaces '#Date#' placeholder with actual date
        filename = w46_template_fn.replace('#IOT#', nameOfExcelFile)
#        filename = filename.replace('#Date#', str(today))
        # wb.save(filename)
        df_final.to_csv(W46_saveToRaw+'W46A_Contractor_Matches_'+nameOfExcelFile+'.csv', encoding = 'utf-8', index = False)
        wb.save(W46_saveTo + filename)
        
   except IllegalCharacterError: 
        logger.error("Illegal Character Error found while creating excel file, solution: Escaping Characters") 
        
 #       print(df_final['Contractor CNUM'])
        df_final = df_final.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
        
        for r in dataframe_to_rows(df_final, index=False, header=True):
            ws.append(r)       
        
        #Saves the file with new name and replaces '#Date#' placeholder with actual date
        filename = w46_template_fn.replace('#IOT#', nameOfExcelFile)
#        filename = w46_template_fn.replace('#Date#', str(today))        
        # wb.save(filename)
        df_final.to_csv(W46_saveToRaw+'W46A_Contractor_Matches_'+nameOfExcelFile+'.csv', encoding = 'utf-8', index = False)
        wb.save(W46_saveTo + filename)
        
   except SyntaxError:
        logger.error('Exception in process_df_to_excel:', exc_info=True)

#%%
def validate_report(fpath):
   """ A function to validate the input file
   Parameters
   ----------
   fpath: str
       a valid path to the file
  
   Returns
   ---------
       a boolean True/False
   """
   try:
    
        if os.path.isfile(fpath) and os.path.getsize(fpath) > 0:
            
            with open(fpath,errors='ignore') as f:
                
                row_count =  sum(1 for line in f)
            
            if row_count > 1:
                return True
            else:
                return False
        else:
            return False  
        
   except SyntaxError:
        logger.error('Exception in should_I_run():', exc_info=True)

#%%
def getContractors(worklocVal = workloc, fileNameVal = filename_cld ):
    """ Main function to extract w46 data from wex-api 
    
    Parameters
    ----------
    worklocVal: str, optional
        a valid workLocation like 'Europe', 'Japan',etc.
        if not added it will load default value from configuration ini file
    fileNameVal: str, optional
        a full path and file name to cld file.
        if not added it will load default value from configuration ini file
    
    Returns
    ---------
    excel
        final w46 report for specific workLocation
       
    """
    
    #Recording start time
    start = time.perf_counter()
    demand_df =  pd.DataFrame()
    practNoMatches = pd.DataFrame()
    df_final_Local = pd.DataFrame()
    temp = pd.DataFrame()

    if validate_report(fileNameVal):
        practNoMatches = createNoMatches_df()  
        demand_df = create_demand(fileNameVal, worklocVal)
        payload = createPayload(worklocVal)
        logger.info("\tProcessing for:{} \tTotal demand records:{}".format(worklocVal,len(demand_df)))
        # logger.info("\tTotal demand records:{}".format(len(demand_df)))
        
        print(f"\nTotal records for {worklocVal}: {len(demand_df)} ")    

        
        #Adding progress bar tqdm library        
        prog_bar = tqdm(demand_df.index,total=len(demand_df), unit='per Seat', unit_divisor=1024)        
        cnt =0
        for i in prog_bar:
            json = get_API_Data(i, payload,demand_df)
            if json['status_code'] == 200:
                temp = process_API_data(i,json,demand_df,practNoMatches)
                df_final_Local = pd.concat([df_final_Local,temp])               
                prog_bar.set_description("Serial:{}".format(i))
                cnt = cnt+1



        #finish time for API run
        finish = time.perf_counter()
        logger.info('API finshed in {} seconds(s)'.format(round(finish-start,2)))     
        

        if df_final_Local.empty:
            finish = time.perf_counter()
            logger.info('Dataframe is empty: Halted at {} seconds(s)\n'.format(round(finish-start,2)))
            raise AttributeError('Dataframe is empty')

        else:
            df_final_Local = format_final_df(df_final_Local)
   #         df_final_Local.to_excel('df_final_local.xlsx', encoding = 'utf-8', index = True)
            process_df_to_excel(worklocVal, df_final_Local)
            print("Done!")
        
    else:
        print("\nFile for {} is empty or does not exists".format(worklocVal))
        logger.error("File for {} is empty or does not exists:{}".format(worklocVal, fileNameVal))
    #Records elapsed time for overall process   
    finish = time.perf_counter()
    logger.info('Overall in {} seconds(s)\n'.format(round(finish-start,2)))
    
#%%
# if __name__ == '__main__':
#     df_final = pd.DataFrame()
# main(workloc)
# getContractors(worklocVal ='Asia Pacific',fileNameVal = 'W46_CLD_extraction8small.csv') # overwrites value from ini, using default filename
# getContractors(fileNameVal = 'W46_CLD_extraction_small.csv') #using default values from ini file
# files_path = "C:\\1Projects\\Enhanced Matching\\codes\\ContractorReplacement\\v4\\CLD_Extraction_Files\\"
# getContractors(worklocVal='Europe',fileNameVal = files_path + 'W46_CLD_extraction_GC.csv')