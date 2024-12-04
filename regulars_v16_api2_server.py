# -*- coding: utf-8 -*
"""    teste
Created on Thu Mar  5 16:36:14 2020
@author: MINORGOMEZVIALES
Version created on February 1st
@autor: Edinelson Almeida
W46 enhance matching...
The purpose of this code is to look up for IBM Regular Practitioners who are in Bench or will be in Bench within a pre-set number of days and find potential 
Demand matches that these IBMers could be assigned to.
The source of data is the Professional Marketplace (ProM) application, for both Practitioners and Open Seats.
This code uses WEX API ENHANCE MATCHING to fetch all the possible available matches.
"""
#Modules for pandas and numpy
from typing import final
import pandas as pd
import numpy as np

#(new) Modules for round
import math

#(new) Module for iteration
import itertools


#Modules for connections
import requests
import urllib3

#Modules for logger
import logging, logging.handlers
import logging

#Modules for dates and time
import time
import datetime
from datetime import timedelta
from datetime import date
from timeit import default_timer as timer
    
#Module for excel
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows

#Module for config.ini
import configparser
from configparser import ConfigParser, ExtendedInterpolation
import ast

#Module for progress bar
#from tqdm import tqdm

import json
import sys
import os

#import pprint
# import os
# import threading
# import concurrent.futures
#%%

class Error(Exception):
    """Base class for other exceptions"""
    pass

class IOTNotValid(Error):
    """Raised when the IOT in parameter is no valid"""
    pass
class IOTNotValidNull(Error):
    """Raised when the IOT in parameter is no valid"""
    pass

class TypeError_500(Error):
    """Raised when POST returns error 500"""
    pass
    
class EmptyPractitioner(Error):
    """Raised when JSON returned from API is empty"""
    pass
    
    
    
#Global variables
try:
    #Disable warnings
    urllib3.disable_warnings()
    
    ######
    IOTS = ['Americas','EMEA','APAC','GCG','JP']
    IOT_par = ''
    if len(sys.argv) == 1:
       raise IOTNotValidNull

    IOT_par = sys.argv[1]
    if IOT_par not in IOTS:       
       raise IOTNotValid
    
    #Parser *.ini file configuration
    config_normal = configparser.ConfigParser()
    configInterpolation = ConfigParser(interpolation=ExtendedInterpolation())
    configInterpolation.read('configRegulars_v6_api2_test.ini')
    config_normal.read('configRegulars_v6_api2_test.ini')
    countryPI = ast.literal_eval(config_normal.get('countriesPI','countryPI'))    

    
    err500_counter = 0
    zero_matches_counter = 0
    seat_counter = 0
    skipped_30_Days_Counter = 0
    fieldsToRetrieve_OS = ""
    match_fields_ret =""
    
    #BEGIN log configuration
    LOG_FILENAME = "logs\\GBSRegulars_" + IOT_par + "_" +str(datetime.datetime.utcnow().strftime('%Y_%m_%d_%I_%M_%S')) + '.log'
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)    
    formatter = logging.Formatter('%(asctime)s:%(levelname)s:%(name)s: %(message)s')    
    file_handler = logging.FileHandler(LOG_FILENAME)
    file_handler.setFormatter(formatter)    
    stream_handler = logging.StreamHandler()
    # stream_handler.setFormatter(formatter)    
    logger.addHandler(file_handler)
    # logger.addHandler(stream_handler)
    logger.info('IOT: '+IOT_par+ '. Processing started at ' + str(datetime.datetime.now()))
    
   
except configparser.ParsingError as e:
    print("Error reading the file:")
    print(e)
    logger.error('ParsingError loading config file:', exc_info=True)
except configparser.NoSectionError as e:
    print("Secified section is not found:")
    print(e)
    logger.error('NoSectionError loading config file:', exc_info=True)
except configparser.NoOptionError as e:
    print("Specified option is not found in the specified section:")
    print(e)
    logger.error('NoOptionError loading config file:', exc_info=True)    
except configparser.Error as e:
    print("Error: ")
    print(e)
    logger.error('Exception loading config file:', exc_info=True)
except IOTNotValid as e:
    print("IOT "+IOT_par +" is Not valid!")
#    print(e)
    logger.info(IOT_par + ' is an invalid IOT.', exc_info=True)
    sys.exit("Enter a valid IOT")  
except IOTNotValidNull as e:
    print("Parameter IOT is Null!")
#    print(e)
    logger.info("Parmeter IOT is Null!. Pass a valid IOT.", exc_info=True)
    sys.exit("Enter a valid IOT") 

#%%   
def add_days_from_current_date (days):
    """
    created by Edinelson Almeida
    This function converts int to unix format date
    parameter: days, example: 120 (days)
    This is required as a filter in the payload for:
        AVAILABILITY_DATE_SECONDS
    """
    new_date = datetime.datetime.now() + timedelta(days=int(days))
    d = datetime.date(new_date.year,new_date.month,new_date.day)
    new_date_unix_format = time.mktime(d.timetuple())
    return new_date_unix_format   


#%%  
def matchingRules():
    global fieldsToRetrieve_OS, match_fields_ret
    try:
    
        Practworklocation_setting = ""
        IOT_name = sys.argv[1]
        
        if IOT_name == 'Americas':
            Practworklocation_setting = "Americas"
        elif IOT_name == 'APAC':
            Practworklocation_setting =  "APAC"         
        elif IOT_name == 'EMEA':
            Practworklocation_setting = "EMEA"
#        elif IOT_name == 'GCG':
#            Practworklocation_setting = "APAC"
        elif IOT_name == 'JP':
            Practworklocation_setting = "Japan"
            

        #basic parameters
        availabilitydays            = config_normal.get('basic_params','availabilitydays')
        PractLob_setting            = config_normal.get('basic_params','lob')
        #Practworklocation_setting   = config_normal.get('basic_params','workloc')
        PractinitialDate_setting_P  = config_normal.get('basic_params','initialdat_p')
        PractfinalDate_setting_P    = config_normal.get('basic_params','finaldate_p')
        PractinitialDate_setting_OS = config_normal.get('basic_params','initialdat_os')
        PractfinalDate_setting_OS   = config_normal.get('basic_params','finaldate_os')
        # talent_pool_type            = config_normal.get('basic_params','talentPoolType')    
        numReqRecs                  = config_normal.get('basic_params','numReqRecs')
        numReqRecsPostSorting       = config_normal.get('basic_params','numReqRecsPostSorting')
        fieldstoretrieve_pract      = ast.literal_eval(config_normal.get ('colsConfig','fieldstoretrieve_pract')) 
        fieldsToRetrieve_OS         = ast.literal_eval(config_normal.get('colsConfig','fieldstoretrieve_os'))
        match_fields_ret            = ast.literal_eval(config_normal.get('colsConfig','match_fields'))
        
        
        #(new) adding parameters for multiple API Calls
        PractintervalpartsDate_setting_P    = config_normal.get('basic_params','intervalpartsdate_p')
        PractintervalinitialDate_setting_P    = config_normal.get('basic_params','intervalinitialdat_p')
        Practintervalpastdays_setting_P = config_normal.get('basic_params','intervalpastdays_p')
        
        #process dates variables
        INTERVAL_START_OS = add_days_from_current_date(PractfinalDate_setting_OS)
        INTERVAL_END_OS = add_days_from_current_date(PractinitialDate_setting_OS)
        INTERVAL_START_P = add_days_from_current_date(PractinitialDate_setting_P)
        INTERVAL_END_P = add_days_from_current_date(PractfinalDate_setting_P)
        
        
        #(new) interval variables for API Calls
        INTERVAL_INITIAL_P=add_days_from_current_date(PractintervalinitialDate_setting_P)
        INTERVAL_PARTS_P = add_days_from_current_date(PractintervalpartsDate_setting_P)-INTERVAL_INITIAL_P
        #(new) interval for past days
        INTERVAL_PASTDAYS = add_days_from_current_date(Practintervalpastdays_setting_P)-INTERVAL_INITIAL_P
       
    
        #(new) interval for last year
        INTERVAL_INITIAL_LASTYEAR=(math.ceil(INTERVAL_INITIAL_P-31536000))

        #(new) interval range for API Calls
        RANGE_INTERVAL_PAST=int(math.ceil((INTERVAL_INITIAL_P-INTERVAL_INITIAL_LASTYEAR)/INTERVAL_PASTDAYS))
        RANGE_INTERVAL=int(math.ceil((int(PractfinalDate_setting_P))/int(PractintervalpartsDate_setting_P)))+RANGE_INTERVAL_PAST+1
        
        #(new) setting LAST_SELECTION as 0 to search all practitioners to date
        LAST_SELECTION=0.0

        #(new) setting variables for Divisions (list of divisions)
        n=5
        lst_remove=[]
        SELECTIONS = [list(i) for i in itertools.product([0, 1], repeat=n)]
        [lst_remove.append(i) if i[3]==0 and i[4]==0 else True for i in SELECTIONS]
        [SELECTIONS.remove(k) for k in lst_remove]
        
        '''
        #(new) Divisions
        if IOT_name=="APAC":
            WORK_LOCATION_COUNTRY_SELECTION=["India", "! India"]
            BAND_SELECTION=["7","! 7"]
            SECTOR_SELECTION=["Cross Sector", "! Cross Sector"]
            JOBROLE_SELECTION=["Application Developer", "! Application Developer"]
            JOBROLE2_SELECTION=["Process Delivery Specialist", "! Process Delivery Specialist"]

        else:
        WORK_LOCATION_COUNTRY_SELECTION="India"
        BAND_SELECTION="!"
        SECTOR_SELECTION="!"
        JOBROLE_SELECTION="!"
        JOBROLE2_SELECTION="!"
        SELECTIONS=[[0,0,0,0]]
        
        DIVISIONS=len(SELECTIONS)
        #print(SELECTIONS)
        '''

        #(new) setting array for matchingRules
        array_matchingRules=[]

        #(new) adding toggle_break to break API Calls if needed
        toggle_break=False
        '''

        for count,division in enumerate(SELECTIONS):
            set_past_days=1; set_last_year=1; set_today=0; set_interval_parts=0; reset_range=0; set_end=0
            LAST_SELECTION=0.0
            toggle_break=False

            SELECTION_COUNTRY=division[0]; SELECTION_BAND=division[1]; SELECTION_SECTOR=division[2]; SELECTION_JOBROLE=division[3]; SELECTION_JOBROLE2=division[4]
        
'''
            

        for range_number in range(RANGE_INTERVAL+1):
                

                #(new) first check for API Call intervals (days before today)
                if (INTERVAL_INITIAL_LASTYEAR+(range_number)*INTERVAL_PASTDAYS<INTERVAL_INITIAL_P):
                    set_past_days=1; set_last_year=1; set_today=0; set_interval_parts=0; factor_past_days=range_number; reset_range=0; set_end=0

                #(new) second check for API Call intervals (days after today)
                else:
    
                    if INTERVAL_INITIAL_P+(range_number-reset_range-1)*INTERVAL_PARTS_P*set_interval_parts>INTERVAL_END_P:
                        set_end=1; set_interval_parts=0; set_today=0; set_past_days=0; set_last_year=0; factor_past_days=0; reset_range=0; toggle_break=True

                        
                    else:
                        set_past_days=0; set_last_year=0; set_today=1; set_interval_parts=1; reset_range=factor_past_days; set_end=0

                
                matchingRules = {"numRecords": numReqRecs,
                                "numRecordsPostSorting": numReqRecsPostSorting,
                                "sourceDocumentType": "PRACTITIONER",
                                "sourceDocumentFields": fieldstoretrieve_pract,                     
                                "sourceDocumentSelection": [
                                    {"name":"WORK_LOCATION_IOT","selections":[Practworklocation_setting]},
                                    #{"name":"MYSA_TALENT_POOL_LOB","selections":[PractLob_setting]},
                                    
                                    #(new) first selection (top limit)
                                    {"name":"AVAILABILITY_DATE_SECONDS","selections":["<= " +str(INTERVAL_INITIAL_LASTYEAR*set_last_year+factor_past_days*INTERVAL_PASTDAYS*set_past_days+INTERVAL_INITIAL_P*set_today+(range_number-reset_range-1)*INTERVAL_PARTS_P*set_interval_parts+INTERVAL_END_P*set_end)]},
                                                                      
                                    #(new) second selection (bottom limit)
                                    {"name":"AVAILABILITY_DATE_SECONDS","selections":["> " +str(LAST_SELECTION)]},
                                    #(new) add another selection to divide by country when using APAC data
                                    #{"name": "WORK_LOCATION_COUNTRY", "selections": ["! India"]},#[WORK_LOCATION_COUNTRY_SELECTION[SELECTION_COUNTRY]]},
                                    {"name": "DEPLOYABILITY", "selections": ["Full"]},
                            
                                    #(new) new selections
#                                    {"name": "BAND", "selections": [BAND_SELECTION[SELECTION_BAND]]},
#                                    {"name": "SECTOR", "selections": [SECTOR_SELECTION[SELECTION_SECTOR]]},
#                                    {"name": "JOB_ROLE", "selections": [JOBROLE_SELECTION[SELECTION_JOBROLE]]},
#                                    {"name": "JOB_ROLE", "selections": [JOBROLE2_SELECTION[SELECTION_JOBROLE2]]},
                            #        {"name":"TALENT_POOL_TYPE","selections":["IBM Regular","Affiliate"]}
                                    ],#End of SourceDocumentSelection
                                "targetDocumentType": "OPEN_SEAT",
                                "targetDocumentFields": fieldsToRetrieve_OS,
                                "targetDocumentSelection":[{"name":"MYSA_SEAT_LOB","selections":[PractLob_setting]},
                                                            {"name":"WORK_LOCATION_COUNTRY", "selections": ["! Austria"]},
                                                            {"name":"CALL_STATUS","selections":["Call"]},
                                                            {"name":"START_DATE", "selections":["< " +str(INTERVAL_START_OS)]},
                                                            {"name":"END_DATE","selections":["> " + str(INTERVAL_END_OS)]},
                                                            {"name": "OPEN_SEAT_ID", "selections": ["! 5140191"]}
                                                            
                                                    ],#End of targetDocumentSelection
                                "rankingRules": {"sortOrder": ["MATCH_SCORE"] ,
                                                 
                                          "matchScoreConfiguration": {
                                            "quantifiers" :["JRS_SCORE", "REQUIRED_SKILLS_SCORE", "BAND_SCORE", "LANGUAGE_SCORE", "LOCATION_SCORE", "AVAILABILITY_SCORE"],
                                            "jrsQuantifiers":{
                                                "PRIMARY_JRS_MATCH": 30.0,	
                                                "SECONDARY_JRS_MATCH":	25.0,
                                                "ADJACENT_JRS_MATCH_100": 25.0,
                                                "ADJACENT_JRS_MATCH_90": 22.0,
                                                "ADJACENT_JRS_MATCH_80": 19.0,
                                                "ADJACENT_JRS_MATCH_70": 16.0,
                                                "ADJACENT_JRS_MATCH_60": 13.0,
                                                "ADJACENT_JRS_MATCH_50": 10.0,
                                                "ADJACENT_JRS_MATCH_40": 7.0,
                                                "ADJACENT_JRS_MATCH_30": 4.0,
                                                "ADJACENT_JRS_MATCH_20": 1.0,
                                                "ADJACENT_JRS_MATCH_10": 0.0,
                                                "NO_JRS_MATCH":	0.0
                                            }, # End of jrsQuantifiers
                                            "requiredSkillsQuantifiers": {
                                                "REQUIRED_SKILLS_MATCH_100": 30.0,
                                                "REQUIRED_SKILLS_MATCH_90": 27.0,
                                                "REQUIRED_SKILLS_MATCH_80": 24.0,
                                                "REQUIRED_SKILLS_MATCH_70": 21.0,
                                                "REQUIRED_SKILLS_MATCH_60": 18.0,
                                                "REQUIRED_SKILLS_MATCH_50": 15.0,
                                                "REQUIRED_SKILLS_MATCH_40": 12.0,
                                                "REQUIRED_SKILLS_MATCH_30": 9.0,
                                                "REQUIRED_SKILLS_MATCH_20": 6.0,
                                                "REQUIRED_SKILLS_MATCH_10": 3.0,
                                                "NO_REQUIRED_SKILLS_MATCH": 0.0,
                                                "REQUIRED_SKILLS_NOT_PROVIDED": 0.0
                                            }, #End of Required SkillsQuantifiers                                  
                                           "bandQuantifiers":{
                                                "BAND_WITHIN_RANGE": 10.0,
                                                "BAND_BELOW_1_RANGE": 5.0,
                                                "BAND_ABOVE_1_RANGE": 0.0,
                                                "NO_BAND_MATCH": 0.0
                                            }, # End of bandQuantifiers
                                            "languageQuantifiers":{
                                                "LANGUAGES_MATCH_FULL": 10.0,
                                                "LANGUAGES_MATCH_PARTIAL": 5.0,
                                                "NO_LANGUAGES_MATCH": 0.0,
                                                "LANGUAGES_NOT_REQUIRED": 0.0
                                            }, # End of languageQuantifiers
                                            "locationQuantifiers":{
                                                "LOCATION_SAME_CITY": 10.0,
                                                "LOCATION_SAME_COUNTRY": 8.0,
                                                "LOCATION_SAME_MARKET": 4.0,
                                                "LOCATION_SAME_GEO_IOT": 2.0,
                                                "NO_LOCATION_MATCH": 0.0
                                            }, # End of locationQuantifiers
                                            "availabilityDateQuantifiers":{
                                                "AVAIL_DATE_WITHIN_7_DAYS_OF_START_DATE": 10.0, 
                                                "AVAIL_DATE_WITHIN_14_DAYS_OF_START_DATE": 5.0,
                                                "AVAIL_DATE_NOT_WITHIN_14_DAYS_RANGE": 0.0,
                                                "PRACTITIONER_AVAILABLE_NOW": 0.0,
                                                "PRACTITIONER_AVAILABLE_WITHIN_7_DAYS": 0.0,
                                                "PRACTITIONER_AVAILABLE_WITHIN_8_14_DAYS": 0.0,
                                                "PRACTITIONER_AVAILABLE_WITHIN_15_21_DAYS": 0.0,
                                                "PRACTITIONER_AVAILABLE_WITHIN_22_30_DAYS": 0.0,
                                                "PRACTITIONER_NOT_AVAILABLE_WITHIN_30_DAYS": 0.0
                                            }, # End of availabilityDateQuantifiers
                                          },
                                        }
                                            
                         }#En of dictionary


                #(new) appending API Calls
                array_matchingRules.append(matchingRules)
                #(new) checking if break needed
                if toggle_break==True:
                    RANGE_INTERVAL=range_number
                    break
                
                #(new) break not needed
                else:
                    True
                
                #(new) setting bottom limit for API Call selection
                LAST_SELECTION=INTERVAL_INITIAL_LASTYEAR*set_last_year+factor_past_days*INTERVAL_PASTDAYS*set_past_days+INTERVAL_INITIAL_P*set_today+(range_number-reset_range-1)*INTERVAL_PARTS_P*set_interval_parts+INTERVAL_END_P*set_end

        # return matchingRules

        #print(array_matchingRules)
        #logger.info(array_matchingRules)
        #print(INTERVAL_INITIAL_LASTYEAR)
        #print(RANGE_INTERVAL_FUTURE)

        #(new) returning API Calls and the number of API Calls
        return array_matchingRules, RANGE_INTERVAL+1#, DIVISIONS
        
    except Exception:
        logger.error('error occurred at function Matching rules:', exc_info=True)
    


#%%
def check_If_AvailDate_Was_Updated (DateField):
   try:
       #Variable to check availability date field
        day = DateField.day
        month = DateField.month
        year = DateField.year         
        availdt =  date(year,month,day).isoformat() 

        PractfinalDate_setting_P   = config_normal.get('basic_params','finaldate_p')
        check_practFinalDate_P  = (datetime.datetime.now() + timedelta(days=int(PractfinalDate_setting_P))).isoformat()[0:10] + ' 12:00:00+00:00'    
                
        if availdt <= check_practFinalDate_P:
            val = True
        else:
            val = False

        return val
   except Exception:
     logger.error('error occurred at function check_If_AvailDate_Was_Updated()', exc_info=True)  
     pass

#%%
#(new) defining get_API_Data with apiRules, APICalls and count
def get_API_Data(apiRules,APICalls,count):
    """
    Connects to API and gets
    takes: cnum from any source
    returns: reads json format and pass to process_API_data() function
    """
    
    #API related parameters
    # base_url = config_normal.get('api','base_url')
    # env      = config_normal.get('api', 'env')
    #uid      = config_normal.get('api', 'uid')
    #pw       = config_normal.get('api', 'pwd')
    #pw       = configInterpolation.get('api', 'pwd')
    url      = config_normal.get('api','getseatforpract_url')
    pw       =  os.getenv('INET_PW') # your intranet id password
    uid      =  os.getenv('INET_UID') # your intranet id password
    
    
    #Local variables
    #---------URL setting to connect the API, search matching seats------------------------
    # search_SeatForPract    = 'getSeatsForPractitioners'
    # url_get_seats = '/'.join([base_url,env,search_SeatForPract])
    # url = url_get_seats
    json = apiRules
    t = 0
    # page_start = time.time()
    start       = time.perf_counter()
    start = timer()
        
    global err500_counter

    try:               
        #(new) Performs connection to the API, printing count and API Calls
        #(new) Reset count when needed
        if count>APICalls:
            count=1
        else:
            True

        print("\nFetching from API ",count,'/',APICalls," ...")
        #(new) Prints to log
        logger.info("Fetching from API " + str(count) + '/' + str(APICalls) + " ...")
        logger.info('Fetching from API...' + str(datetime.datetime.now().time()))    
        print('url=',url)
        
        #(new) time to the server close the connection on the other side
        time.sleep(3)

        #request
        r = requests.post(url,auth=(uid, pw),verify=False,json=json, timeout=7200)
        #print("API Calls remaining:",r.headers["X-Rate-Limit-Remaining"])
        
        
        end2 = timer()
        logger.info('API finished in {} seconds(s) '.format(timedelta(seconds = end2-start)) +' at ' + str(datetime.datetime.now().time()))
        print('API finished in {} seconds(s)'.format(timedelta(seconds = end2-start)))

        #(new) print API Calls remaining to the logs in case we reach the limit
#        print(r.headers.keys())
        logger.info("API Calls remaining:{}".format(str(r.headers.get('X-Rate-Limit-Remaining'))))
        
        if len(r.text) == 0:
            logger.info("Blank response requests error")
        
        if r.status_code == 500:
            err500_counter += 1
            logger.info("Error 500 Internal server")
            raise TypeError ("API returned error")           
#            raise TypeError_500

        if r.status_code == 200:
            print("Loading returned data...")
            logger.info('Loading returned data...'.format(json))
            api_returned_data = r.json()
            print("Counting returned data...")
    
            return api_returned_data
           
    except requests.exceptions.RequestException as e: 
            print('***Exception***')
            print(e)
            logger.error('RequestException at function get_API_Data:', exc_info=True)
    except requests.exceptions.Timeout as e:
            print('***Timeout Exception***')
            print(e)
            logger.error('TimeoutException at function get_API_Data:', exc_info=True)
    except TypeError_500 as e500:
            print("Error 500 Internal server")
            sys.exit("Error 500 Internal server")  
    except Exception as ex:
            print("\nget_API_Data:",ex)
            logger.error('Exception at function get_API_Data:', exc_info=True)
        
#%%
def create_DataFrames(c):

    logger.info('Creating dataframe at:  ' + str(datetime.datetime.now().time()))
        
    global seat_counter, Pract_seatsData_df, df_final, zero_matches_counter,skipped_30_Days_Counter

    try:               
        
        contador1 = 0
        totgeral = 0

        mydict = c
        
        #preparing data frames
        df1 = pd.DataFrame(mydict, columns = ['practitioners']) 
        df2 = pd.DataFrame(mydict, columns = ['openSeats']) 
        df3 = pd.DataFrame(mydict, columns = ['matches'])
       
        isempty_pract = df1.empty 
        if isempty_pract:
           raise EmptyPractitioner
        else:
           prac_df  = df1["practitioners"].apply(pd.Series )
       
        isempty_seat = df2.empty 
        if isempty_seat:
           curr_date  = (datetime.datetime.now()).isoformat()[0:10] + ' 12:00:00+00:00'
           print('There is no OpenSeat')
           list = [ ("None",
                                   "None",
                                   "None",
                                   "None",
                                   curr_date,                                   
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None",
                                   "None" ) ]
           seats_df = pd.DataFrame(list,columns = ["osID",
                                   "title",
                                   "jobRole",
                                   "startDate",
                                   "startDateFormatted",
                                   "endDateFormatted",
                                   "endDate",
                                   "requestedBandLow",
                                   "requestedBandHigh",
                                   "requiredSkills",
                                   "skillSet",
                                   "workLocationIMT",
                                   "workLocationCountry",
                                   "jrs",
                                   "industry","sector",
                                   "contractStatus",
                                   "contractOwningOrganization",
                                   "clientName",
                                   "candidateTrackingStatus",
                                   "acceptContractors",
                                   "backfill",
                                   "backfillType",
                                   "backfillReason",
                                   "ownerName",
                                   "callStatus",
                                   "positionCandidates"])


        else:   
           seats_df = df2["openSeats"].apply(pd.Series )


        isempty_match = df3.empty 
        if isempty_match:
           print('There is no Match')
           matches = pd.DataFrame(columns = ["wexScore", "matchScore","matchScoreQuantifiers","keywordsMatchScore","requiredKeywordsMatchScore","niceToHaveKeywordsMatchScore","matchType","priorityMatchType","jrsMatchType","jrsSimilarityRank","jrsSimilarityScore","jrsSimilarityScorePercentile","projectedTrainingWeeks","topsisScore","externalScore","assignment","openSeat","practitioner"])
        else:
            matches  = df3["matches"].apply(pd.Series )
       
      
        #add sufix
        prac_df = prac_df.add_suffix('_P')
        prac_df.rename(columns={'cnum_P':'cnum'}, inplace = True)    
        seats_df = seats_df.add_suffix('_OS')
        seats_df.rename(columns={'osID_OS':'osID'}, inplace = True)    
#        seats_df.to_csv('seats_df.csv', header=True, index=False, sep='\t', mode='a')

        if isempty_match:
           df4 = pd.DataFrame()
           df5 = pd.DataFrame()
           df4['osID'] = "" 
           df5['cnum'] = ""
        else: 
           df4 = matches["openSeat"].apply(pd.Series )
           df5 = matches["practitioner"].apply(pd.Series)
        
        matches = pd.concat([matches, df4,df5], axis=1)
               
        #remove empty keys
        fil = matches["osID"] != ""
        NewMatches = matches[fil]
        fil = matches["cnum"] != ""
        NewMatches = matches[fil]
        #NewMatches.to_csv('NewMatches.txt', header=True, index=False, sep='\t', mode='a')
        

        #merge 3 dataframes
        if isempty_seat:
           NewMatches = NewMatches[["matchType", "jrsSimilarityScorePercentile","jrsMatchType", "priorityMatchType","matchScore","matchScoreQuantifiers"]]
           merged1 = pd.concat([NewMatches,seats_df], axis =1) 
           merged = pd.concat([merged1, prac_df], axis =1) 
           
        else:
           NewMatches = NewMatches[["osID","cnum","matchType", "jrsSimilarityScorePercentile","jrsMatchType", "priorityMatchType","matchScore","matchScoreQuantifiers"]]
           merged = NewMatches.merge(seats_df, how ='inner', on ='osID')
           merged = merged.merge(prac_df, how = 'right', on = 'cnum') 
 
 
        #remove blank cnums (after right join)
        if not isempty_seat:
           fil = merged["cnum"] != ""
           NewMatches = merged[fil]
           merged = NewMatches

        # remove those ones that had the availability date changed durante the code execution
        if not isempty_seat:
            #temporary column to set whether the record will or not be removed based on availability date changed durante the code execution
            merged['Status_date'] = ''

            for ind in merged.index: 
                availabilityDate = pd.to_datetime(merged['availabilityDate_P'][ind])
                if not check_If_AvailDate_Was_Updated(availabilityDate):
                
                #(new) switched old merged to merged.loc to fix error       
                ##    merged['Status_date'][ind] = 'Remove'
                   merged.loc[ind,'Status_date'] = 'Remove'
                   logger.info(f"CNUM ID Skipped: {merged['cnum']}")
                   skipped_30_Days_Counter += 1 
                else:
                   seat_counter += 1
                
                #(new) switched old merged to merged.loc to fix error 
                #   merged['Status_date'][ind] = 'Keep'
                   merged.loc[ind, 'Status_date'] = 'Keep'
     
            zero_matches_counter = merged.matchScore.isna().sum()
            seat_counter = seat_counter - zero_matches_counter
        
        else:
           merged['Status_date'] = 'Keep'

        merged.drop(merged[merged.Status_date == 'Remove'].index, inplace=True)
        merged.reset_index(drop = True, inplace = True)  
        merged.fillna('None', inplace = True) 
        
        Pract_seatsData_df  = merged
        
        logger.info('Total Practitioner:  ' + str(contador1))
        logger.info('Total general:  ' + str(totgeral))
        
    except EmptyPractitioner as e:
        print("There is no Practitioner returned!")
        logger.info("There is no Practitioner returned!", exc_info=True)
        sys.exit("There is no Practitioner returned!")  
           
#%%
def process_DataFrames(Pract_seatsData_df):
    """
    Process API data and creates dataframe
    takes: json format
    returns: it creates final dataframe
    """

    logger.info('Processing dataframe at: ' + str(datetime.datetime.now().time()))
    
    try:  
        #Columns settings
        #to literally pass in a list then you can use: ast.literal_eval()
        columnsToRename             = ast.literal_eval(config_normal.get('colsConfig','renamecolumns'))
        columnsToDrop               = ast.literal_eval(config_normal.get('colsConfig','dropcolumns'))        

        # 
        finalDataframe  = Pract_seatsData_df
    
        #Change band data due data protection for Italy
        finalDataframe.loc[(finalDataframe.workLocationCountry_P == 'Italy'),'band_P'] = ' '

       #(new) switched old to new to avoid "None" in every row of RSP Name
       #(old)
        # format output of rsaValues 
        #finalDataframe['rsaValues_P'] =  finalDataframe.rsaValues_P.replace("",'None').astype('str')
#        if finalDataframe['rsaValues_P'].notnull().any():
        # if finalDataframe.iloc[0]['rsaValues_P'] == "None":
        #     finalDataframe['rsaValues_P'] = 'None'
        # else:    
        #     test = finalDataframe['rsaValues_P'].str.split("|", n = 5 , expand = True)    
        #     finalDataframe['rsaValues_P'] = test[2]+test[3]+" "+test[4]

        #(new)
        #format output of rsaValues 
        finalDataframe['rsaValues_P'] = finalDataframe.rsaValues_P.replace("",'None').astype('str')
        #change "" for None
        finalDataframe.loc[finalDataframe['rsaValues_P'] == '', "rsaValues_P"] == "None"
        #change values for RSP Names
        finalDataframe.loc[finalDataframe['rsaValues_P'] != 'None', 'rsaValues_P'] = finalDataframe['rsaValues_P'].str.split("|",n=5, expand=True)[2] + finalDataframe['rsaValues_P'].str.split("|",n=5, expand=True)[3] + " " + finalDataframe['rsaValues_P'].str.split("|",n=5, expand=True)[4]


        #Step 2 Create columns    
        # create column for required skills percent
        finalDataframe['requiredSkillsPercent'] = finalDataframe['matchScoreQuantifiers'].astype('str')
        finalDataframe['requiredSkillsPercent'] = finalDataframe['requiredSkillsPercent'].str.split(",",n=6, expand=True)[5]
        for ind in finalDataframe.index: 
             if finalDataframe['requiredSkillsPercent'][ind] == ' NO_REQUIRED_SKILLS_MATCH (0.0)':
                finalDataframe['requiredSkillsPercent'][ind] = '0'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_NOT_PROVIDED (0.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '0'                
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_10 (3.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '10'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_20 (6.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '20'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_30 (9.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '30'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_40 (12.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '40'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_50 (15.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '50'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_60 (18.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '60'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_70 (21.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '70'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_80 (24.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '80'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_90 (27.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '90'
             elif finalDataframe['requiredSkillsPercent'][ind] == ' REQUIRED_SKILLS_MATCH_100 (30.0)':
                 finalDataframe['requiredSkillsPercent'][ind] = '100'

        finalDataframe['Country Matches_M'] = np.where(finalDataframe["workLocationCountry_P"] == finalDataframe['workLocationCountry_OS'],"Yes", "No")        
        finalDataframe['acceptContractors_OS'] = np.where(finalDataframe["acceptContractors_OS"]=='No',"No", "Yes")    
        finalDataframe.loc[(finalDataframe.matchScore == 'None'),'acceptContractors_OS'] = 'None'
        finalDataframe['matchScoreQuantifiers'] = finalDataframe['matchScoreQuantifiers'].astype('str')      

        #Create special filter field  
        finalDataframe['startDateFormatted_OS'] = finalDataframe.startDateFormatted_OS.replace('None',0)
        finalDataframe['startDateFormatted_OS'] = pd.to_datetime(finalDataframe.startDateFormatted_OS)

        # relace blank by None for backfill fields
        finalDataframe['backfill_OS'] =  finalDataframe.backfill_OS.replace(np.nan,'None').astype('str')
        finalDataframe['backfill_OS'] =  finalDataframe.backfill_OS.replace('','None').astype('str')
        finalDataframe['backfillType_OS'] =  finalDataframe.backfillType_OS.replace(np.nan,'None').astype('str')
        finalDataframe['backfillType_OS'] =  finalDataframe.backfillType_OS.replace('','None').astype('str')
        finalDataframe['backfillReason_OS'] =  finalDataframe.backfillReason_OS.replace(np.nan,'None').astype('str')
        finalDataframe['backfillReason_OS'] =  finalDataframe.backfillReason_OS.replace('','None').astype('str')
        
        # IllegalCharacterError - replace Vertical Tabulation by space for requiredSkills_OS field
        finalDataframe['requiredSkills_OS'] = finalDataframe['requiredSkills_OS'].astype('str')
        finalDataframe['requiredSkills_OS'] = finalDataframe['requiredSkills_OS'].str.replace("\u000b", "\u0020", regex=True)
        # IllegalCharacterError - replace STX - Start Of Text by space for requiredSkills_OS field
        finalDataframe['requiredSkills_OS'] = finalDataframe['requiredSkills_OS'].str.replace("\u0002", "\u0020", regex=True)

        #add ' when the first character of the field is = 
        finalDataframe.loc[(finalDataframe.requiredSkills_OS.str[:1] == '=') ,'requiredSkills_OS'] = "'"+finalDataframe['requiredSkills_OS']+"'"
 
        # replace blank by None for serviceArea_P field
        finalDataframe['serviceArea_P'] = finalDataframe['serviceArea_P'].astype('str')      
        finalDataframe['serviceArea_P'] =  finalDataframe.serviceArea_P.replace('','None').astype('str')

        #converts 'availabilityDate' to date format
        finalDataframe['availabilityDate_P'] = pd.to_datetime(finalDataframe.availabilityDate_P, utc=True)
        
        #Declare variables for bench text calculation
        today   = datetime.datetime.now().isoformat()[0:10] + ' 12:00:00+00:00'
        d7  = (datetime.datetime.now() + timedelta(days=int(7))).isoformat()[0:10] + ' 12:00:00+00:00'
        d14 = (datetime.datetime.now() + timedelta(days=int(14))).isoformat()[0:10] + ' 12:00:00+00:00'
        d21 = (datetime.datetime.now() + timedelta(days=int(21))).isoformat()[0:10] + ' 12:00:00+00:00'
        d30 = (datetime.datetime.now() + timedelta(days=int(30))).isoformat()[0:10] + ' 12:00:00+00:00'
 
        # Create new column Os_Start_Dates_M based on startDateFormated_OS
        conditionsStartDatesM =[
            (finalDataframe.startDateFormatted_OS <= d7),
            (finalDataframe.startDateFormatted_OS <= d14),
            (finalDataframe.startDateFormatted_OS <= d21)     
            ]
        choices1 =["Start Date in <= 7 days",
                   "Start Date in 8-14 days", 
                   "Start Date in 15-21 days"]
        
        finalDataframe['OS_Start_Dates_M']= np.select(conditionsStartDatesM,choices1,default ="Start Date in 22-30 days") 

        #Create column based on availability date
        conditionsAvlMtchM =[
            (finalDataframe.availabilityDate_P < today),
            (finalDataframe.availabilityDate_P <= d7),
            (finalDataframe.availabilityDate_P <= d14),
            (finalDataframe.availabilityDate_P <= d21),
            (finalDataframe.availabilityDate_P <= d30)
            ]
        choices2 =["Already on Bench", 
                   "Bench in 1-7 days",
                   "Bench in 8-14 days",
                   "Bench in 15-21 days",
                   "Bench in 22-30 days"]
        
        finalDataframe['AvailMatch_M']= np.select(conditionsAvlMtchM,choices2, default = "Bench in more than 30 days") 

        finalDataframe['Template Version']= 1

        finalDataframe['Date of Data Retrieval']=datetime.datetime.now().isoformat()[0:10] 
     
        

        #split position candidates and select the CNUM of candidate and the Status/postition 
        for ind in finalDataframe.index: 
           position_list = [] 
           position_str =""
           conector = ""
           for position_cand in finalDataframe['positionCandidates_OS'][ind]:
               pos_split = position_cand.split("|",3)
               if len(pos_split) > 1:
                  position_list.append(pos_split[0]+' - '+pos_split[2])
#                  position_list.append(pos_split[0]+' - '+pos_split[2]+'\r\n')
                  position_str = position_str + conector + pos_split[0]+' - '+pos_split[2] 
                  conector = " || "
                  
                  #",".join(position_list)

           
           finalDataframe.loc[ind,'positionCandidates_OS'] = position_str
#           finalDataframe['positionCandidates_OS'][ind] = position_str
           
        #replace [] by blank for positionCandidates_OS field
        finalDataframe['positionCandidates_OS'] = finalDataframe['positionCandidates_OS'].astype('str')      
        finalDataframe['positionCandidates_OS'] = finalDataframe['positionCandidates_OS'].str.replace("'", "")
        finalDataframe['positionCandidates_OS'] = finalDataframe['positionCandidates_OS'].str.replace("[", "", regex=True)
        finalDataframe['positionCandidates_OS'] = finalDataframe['positionCandidates_OS'].str.replace("]", "", regex=True)
        finalDataframe['positionCandidates_OS'] = finalDataframe.positionCandidates_OS.replace('','None').astype('str')
        
        #replace [] by blank for skills_P field
        finalDataframe['skills_P'] = finalDataframe['skills_P'].astype('str')      
        finalDataframe['skills_P'] = finalDataframe['skills_P'].str.replace("'", "")
        finalDataframe['skills_P'] = finalDataframe['skills_P'].str.replace("[", "", regex=True)
        finalDataframe['skills_P'] = finalDataframe['skills_P'].str.replace("]", "", regex=True)
        finalDataframe['skills_P'] = finalDataframe.skills_P.replace('','None').astype('str')
        
        
        finalDataframe.loc[(finalDataframe.matchScore == 'None'),'matchScore'] = '0'
        finalDataframe['matchScore'] = pd.to_numeric(finalDataframe['matchScore'])
        
       

        #Step 1 Create index
        finalDataframe.rename(columns={'cnum':'Serial Number'}, inplace = True)    
        finalDataframe.set_index('Serial Number', inplace = True)   
              
        #Step 3 Sort columns
        #sort 'cnum','jrsMatchType' - ascending order
        finalDataframe.sort_index(inplace=True)
        finalDataframe.sort_values(by=['jrsMatchType'], inplace=True)    
        # sort qualityScore_OS - descending order
        finalDataframe.sort_values(by=['matchScore'], inplace=True, ascending=False)

        finalDataframe['matchScore'] = finalDataframe['matchScore'].apply(str)
        finalDataframe.loc[(finalDataframe.matchScore == '0.0'),'matchScore'] = 'None'

                
        # Step 4. Converting date columns to values: P_Avail_Date,D_Start_Date, D_End_Date and startDateFormated_OS.
        finalDataframe['availabilityDate_P'] =  finalDataframe.availabilityDate_P.replace('None',0)
        finalDataframe['startDateFormatted_OS'] =  finalDataframe.startDateFormatted_OS.replace('None',0)
        finalDataframe['endDateFormatted_OS'] =  finalDataframe.endDateFormatted_OS.replace('None',0)
        finalDataframe['availabilityDate_P']      = pd.to_datetime(finalDataframe['availabilityDate_P']).dt.date
        finalDataframe['startDateFormatted_OS']   = pd.to_datetime(finalDataframe['startDateFormatted_OS']).dt.date
        
        #(old)
        #finalDataframe['endDateFormatted_OS']     = pd.to_datetime(finalDataframe['endDateFormatted_OS']).dt.date

        #(new) Changed to new because of datetime limitation ("Out of Bounds error")
        #Change to indicator '1678-01-01 00:00:00 GMT'
        #finalDataframe['startDateFormatted_OS'] =  finalDataframe.startDateFormatted_OS.replace(0,'1678-01-01 00:00:00 GMT')
        try:
            finalDataframe.loc[finalDataframe['endDateFormatted_OS'] != 'None', 'endDateFormatted_OS']=np.array(finalDataframe['endDateFormatted_OS'].str[0:10], dtype='str')
            finalDataframe['endDateFormatted_OS'] = finalDataframe['endDateFormatted_OS'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d').date())
        except:
            True
        finalDataframe.loc[(finalDataframe.matchScore == 'None'),'startDateFormatted_OS'] = 'None'
        finalDataframe.loc[(finalDataframe.matchScore == 'None'),'OS_Start_Dates_M'] = 'None'

        #(new)
        finalDataframe.loc[(finalDataframe.matchScore == 'None'),'endDateFormatted_OS'] = 'None'
        
        #Step 5. Drop columns from DataFrame
        finalDataframe.drop(columnsToDrop, axis = 1, inplace = True) 
        
        #Step 6. rename columns
        finalDataframe.rename(columns=columnsToRename, inplace = True) 

        #(new) Step7. dropping duplicate rows remaining from JSON consolidation
        # #drop duplicates
        finalDataframe.drop_duplicates(keep='last',inplace=True)

        #(new) Step 8. dropping duplicate rows with same openseat and notesID
        # #drop duplicates by same openseat and notesID
        finalDataframe.drop_duplicates(subset=['Open Seat ID','Notes ID'],keep='last',inplace=True)

        #(new) Step 9. changing "None" to 0
        #change none to 0.0 to convert to float
        finalDataframe.loc[(finalDataframe['Match Score'] == 'None'),'Match Score'] = '0.0'
        
        #(new) Step 10. converting Match Score column to float
        #Match Score column to float
        finalDataframe['Match Score'] = finalDataframe['Match Score'].apply(float)

        #(new) sorting top 20 records to avoid more than 20 items after JSON consolidation
        #sort top 20 records by match score and group by name
        finalDataframe = finalDataframe.sort_values(['Match Score'],ascending = False).groupby('Notes ID').head(20)
        
        #(new) converting Match Score to string again
        #return Match Score to string
        finalDataframe['Match Score'] = finalDataframe['Match Score'].apply(str)
        
        #(new) changing 0 to "None" again
        #return 0.0 to none
        finalDataframe.loc[(finalDataframe['Match Score'] == '0.0'),'Match Score'] = 'None'

        
        #Saving dataframes
        finalDataframe.fillna('None', inplace = True) 
#        finalDataframe.to_excel('finaldataframe.xlsx', encoding = 'utf-8', index = True)

        #return finalDataframe after processing
        return finalDataframe

    except Exception as ex:
            print("\nprocess_DataFrames:",ex)
            logger.error('error occurred at function process_DataFrames:', exc_info=True)

#%%
def process_df_to_excel(finalDataframe):
    """
    creates excel file using the final dataframe
    takes: dataFrame, local variables, matchingRules dataFrame
    returns: saves xls file
    """
    logger.info('Dataframe to Excel at:' + str(datetime.datetime.now().time()))
     
    try:
        
        #Excel parameters
        # wrk_dir              = config_normal.get('excel','wrk_dir')
        w46_template_fn      = config_normal.get('excel','w46_template')
        W46_source_worksheet = config_normal.get('excel','W46_source_worksheet')
        W46_saveTo          = config_normal.get('excel','w46_excel_saveTo')
        W46_template_location = config_normal.get('excel','w46_template_path')
        W46_Demand_worksheet = config_normal.get('excel','W46_Demand_worksheet')
        W46_saveToRaw         = config_normal.get('excel','wrk_dir_raw')
        
        today   = datetime.datetime.now().isoformat()[0:10]
        #change to working directory 
        # os.chdir(wrk_dir)

        #Reset index
        finalDataframe = finalDataframe.reset_index()
        
        # Removes column 'index' as it only contains '0's
        if 'index' in finalDataframe.columns:
            #source1_df = source1_df
            finalDataframe.drop('index',axis = 1, inplace=True)            
      

        #Points to template
        fn =  W46_template_location + w46_template_fn 
        #load_workbook in memory to open the template and assigned it to 'wb' variable
        wb = xl.load_workbook(filename = fn ,data_only=False, keep_vba=True)
        #Assigns 'about' worksheet to variable
        ws=[W46_Demand_worksheet]

        #Assigns 'Source1' worksheet to variable ('ws')
        ws=wb[W46_source_worksheet]        
        #Removes old sheet 'Source1' from main template workbook (wb)
        wb.remove(ws)       
        #Adds 'Source1' tab and set it as the active sheet
        ws = wb.create_sheet(W46_source_worksheet)
        #transfer the 'source1_df' to the active sheet using append(), removes index and keeps the headers
        
        for r in dataframe_to_rows(finalDataframe, index=False, header=True):
            ws.append(r)
       
        #Saves the file with new name and replaces '#Date#' placeholder with actual date
        # filename = W46_template_location + fn.replace('#Date#',today)
        nameOfIOT = sys.argv[1]
        filename = w46_template_fn.replace('#IOT#', nameOfIOT)
        filename = filename.replace('#Date#',str(today))
        filename = filename.replace('#LOB#',config_normal.get('basic_params','lob'))
        wb.save(W46_saveTo + filename)
        
        if nameOfIOT == 'EMEA':
           finalDataframe.to_csv(W46_saveToRaw+'W46A_Regular_Matches_'+nameOfIOT+'_PI.csv', encoding = 'utf-8', index = False)

           #Change data due data protection for Europe
           finalDataframe.loc[finalDataframe['Country_P'].isin(countryPI),'Notes ID'] = 'Masked'
           finalDataframe.loc[finalDataframe['Country_P'].isin(countryPI),'Intranet_ID_P'] = 'Masked'
           finalDataframe.loc[finalDataframe['Country_P'].isin(countryPI),'Serial Number'] = 'Masked'
           finalDataframe.loc[finalDataframe['Country_P'].isin(countryPI),'Name_P'] = 'Masked'

           finalDataframe.to_csv(W46_saveToRaw+'W46A_Regular_Matches_'+nameOfIOT+'_NON_PI.csv', encoding = 'utf-8', index = False)
        else:
           finalDataframe.to_csv(W46_saveToRaw+'W46A_Regular_Matches_'+nameOfIOT+'.csv', encoding = 'utf-8', index = False)
            
    except Exception as ex:
        print("process_df_to_excel Error:")
        print(ex)
        logger.error('error occurred at function process_df_to_excel:', exc_info=True)
#%%
def main():
    print("\nStarting...")
    
    #Recording start time
    start       = time.perf_counter()
    df_final    = pd.DataFrame()
    apiDataFrame = pd.DataFrame()

    #(new) using first item (0, the rulesforAPI) of the array
    rulesForAPI  = matchingRules()[0]

    #(new) using second item (1, the number of APICalls) of the array
    #(new) Using APAC exception
    #if IOT_par=="APAC":
    #    APICalls=matchingRules()[1]*matchingRules()[2]
    #else:
    APICalls=matchingRules()[1]

    #(new) creating the array to receive the API responses
    #creating array to receive API responses
    apiReturnedArray=[]

    #(new) creating the consolidated variables
    #creating consolidated variables for JSONs
    consolidated_numberOfMatches=0
    consolidated_numberOfDocumentsWithoutMatches=0
    consolidated_totalMatchesPerDocument=[]
    consolidated_practitioners=[]
    consolidated_openSeats=[]
    consolidated_matches=[]



    logger.info('Preparing payload at: ' + str(datetime.datetime.now().time()))    
    print("\nPreparing payload...")

    with open('payload.json', 'w', encoding='utf-8') as f:
        json.dump(rulesForAPI, f,ensure_ascii=False, indent=4)
    
    start = timer()
    
    print("\nGetting data from API...")

    #(new) number of API Calls
    print("\nNumber of API Calls:", APICalls)
    
    #(new) appending API responses
    for count,i in enumerate(rulesForAPI):
        apiReturnedArray.append(get_API_Data(i,APICalls,count+1))
        # apiDataFrame = get_API_Data(rulesForAPI)

    #(new) creating a dictionary and consolidating JSONs
    for i in range(len(apiReturnedArray)):
        
        consolidated_numberOfMatches=str(int(consolidated_numberOfMatches)+int(apiReturnedArray[i]["numberOfMatches"]))
        consolidated_numberOfDocumentsWithoutMatches=str(int(consolidated_numberOfDocumentsWithoutMatches)+int(apiReturnedArray[i]["numberOfDocumentsWithoutMatches"]))
        # consolidated_totalMatchesPerDocument.append(apiReturnedArray[i]["totalMatchesPerDocument"])
        consolidated_totalMatchesPerDocument={**apiReturnedArray[i-1]["totalMatchesPerDocument"],
                                              **apiReturnedArray[i]["totalMatchesPerDocument"]}
        consolidated_practitioners.extend(apiReturnedArray[i]["practitioners"])
        consolidated_openSeats.extend(apiReturnedArray[i]["openSeats"])                    
        consolidated_matches.extend(apiReturnedArray[i]["matches"])
      

    jsonConsolidatedDict={"numberOfMatches": consolidated_numberOfMatches,
                          "numberOfDocumentsWithoutMatches": consolidated_numberOfDocumentsWithoutMatches,
                          "totalMatchesPerDocument":consolidated_totalMatchesPerDocument,
                          "practitioners":consolidated_practitioners,
                          "openSeats":consolidated_openSeats,
                          "matches":consolidated_matches}
    
    

    #(new) passing values from API responses (JSON consolidated) to variable
    apiDataFrame=jsonConsolidatedDict

 

    print("\nPreparing output file...")
    logger.info('Preparing outputfile at:  ' + str(datetime.datetime.now().time()))    

    with open('outputfile.json', 'w') as file:
        file.write(json.dumps(apiDataFrame, indent=4))
        
    end2 = timer()
    logger.info('Data preparation finshed at: ' + str(datetime.datetime.now().time()))    
       
    f = open('outputfile.json') 
    apiDataFrame = json.load(f)

    
    # Closing file 
    f.close()
    
    print("\nCreating Dataframes...")
    create_DataFrames(apiDataFrame)
   
    # Call function to format dataframe  
    print("\nProcessing Dataframes...")
    df_final = process_DataFrames(Pract_seatsData_df)    
#   
 
    if df_final is None:
        logger.info(f"\nTotal Records: {seat_counter}\nNo matches: {zero_matches_counter}\nRecords null:{err500_counter}\nTotal Skipped:{skipped_30_Days_Counter}\nDataframe rows:0\n")
        raise AttributeError("DataFrame is empty")
    else:
        print("\nCreating excel report...")
        process_df_to_excel(df_final)
        #Records elapsed time for overall process 
        logger.info(f"\nTotal Records: {seat_counter}\nNo matches: {zero_matches_counter}\nRecords null:{err500_counter}\nTotal Skipped:{skipped_30_Days_Counter}\nDataframe rows:{len(df_final)}\n")

    #convert seconds to hours, minutes and seconds
    def converter_seconds (seconds):
        hours=seconds//3600
        minutes = (seconds % 3600)//60
        seconds_final = (seconds % 3600) % 60

        return "%0d hours : %0d minutes : %0d seconds" % (hours, minutes, seconds_final)
    
    finish = timer()
    #logger.info('Overall in {} seconds(s)  '.format(round(finish-start,2)) +" at " +str(datetime.datetime.now().time()))
    
    #(new) converted time
    logger.info('Overall in {} '.format(converter_seconds(round(finish-start,2)) +" at " +str(datetime.datetime.now().time())))
    print(f"\nTotal Records: {seat_counter}\nNo matches: {zero_matches_counter}\nRecords null:{err500_counter}\nTotal Skipped:{skipped_30_Days_Counter}\nDataframe rows:{len(df_final)}\n")  
    
#%%
if __name__ == '__main__':
   main()